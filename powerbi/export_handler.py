"""
Power BI Export Handler
Exports query results to Excel/CSV and generates DAX measures and
Power BI setup instructions via Claude API.
"""

import io
import json
from pathlib import Path
from typing import Any

import anthropic
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from config import ANTHROPIC_API_KEY, CLAUDE_MODEL


# ─────────────────────────────────────────────
# Excel helpers
# ─────────────────────────────────────────────

def _style_header_row(ws, num_cols: int) -> None:
    """Apply bold, blue header styling to the first row."""
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    thin = Side(style="thin", color="FFFFFF")
    border = Border(left=thin, right=thin, bottom=thin, top=thin)

    for col_idx in range(1, num_cols + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border


def _auto_column_width(ws) -> None:
    """Set column widths based on content length."""
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                max_len = max(max_len, len(str(cell.value or "")))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max(max_len + 4, 12), 50)


def _alternate_row_fill(ws, num_rows: int, num_cols: int) -> None:
    """Apply alternating row fill for readability."""
    alt_fill = PatternFill("solid", fgColor="EBF3FB")
    for row_idx in range(2, num_rows + 2):
        if row_idx % 2 == 0:
            for col_idx in range(1, num_cols + 1):
                ws.cell(row=row_idx, column=col_idx).fill = alt_fill


# ─────────────────────────────────────────────
# Export functions
# ─────────────────────────────────────────────

def export_to_excel(results_df: pd.DataFrame, filepath: str | Path) -> Path:
    """
    Export query results to a formatted .xlsx file.

    Args:
        results_df: DataFrame of query results.
        filepath:   Destination .xlsx path.

    Returns:
        Path to the written file.
    """
    filepath = Path(filepath)
    filepath.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "Query Results"

    if results_df.empty:
        ws["A1"] = "No data returned."
        wb.save(filepath)
        return filepath

    # Write header
    for col_idx, col_name in enumerate(results_df.columns, start=1):
        ws.cell(row=1, column=col_idx, value=col_name)

    # Write data rows
    for row_idx, row in enumerate(results_df.itertuples(index=False), start=2):
        for col_idx, value in enumerate(row, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    _style_header_row(ws, len(results_df.columns))
    _alternate_row_fill(ws, len(results_df), len(results_df.columns))
    _auto_column_width(ws)
    ws.freeze_panes = "A2"

    wb.save(filepath)
    return filepath


def export_to_excel_bytes(results_df: pd.DataFrame) -> bytes:
    """
    Export query results to Excel and return as bytes (for Streamlit download).
    """
    buf = io.BytesIO()
    wb = Workbook()
    ws = wb.active
    ws.title = "Query Results"

    if not results_df.empty:
        for col_idx, col_name in enumerate(results_df.columns, start=1):
            ws.cell(row=1, column=col_idx, value=col_name)
        for row_idx, row in enumerate(results_df.itertuples(index=False), start=2):
            for col_idx, value in enumerate(row, start=1):
                ws.cell(row=row_idx, column=col_idx, value=value)
        _style_header_row(ws, len(results_df.columns))
        _alternate_row_fill(ws, len(results_df), len(results_df.columns))
        _auto_column_width(ws)
        ws.freeze_panes = "A2"
    else:
        ws["A1"] = "No data returned."

    wb.save(buf)
    return buf.getvalue()


def export_to_csv(results_df: pd.DataFrame, filepath: str | Path) -> Path:
    """
    Export query results to CSV.

    Args:
        results_df: DataFrame of query results.
        filepath:   Destination .csv path.

    Returns:
        Path to the written file.
    """
    filepath = Path(filepath)
    filepath.parent.mkdir(parents=True, exist_ok=True)
    results_df.to_csv(filepath, index=False, encoding="utf-8-sig")
    return filepath


def export_to_csv_bytes(results_df: pd.DataFrame) -> bytes:
    """Return CSV as bytes for Streamlit download."""
    return results_df.to_csv(index=False, encoding="utf-8-sig").encode("utf-8-sig")


def export_standardized_excel(
    results_df: pd.DataFrame,
    viz_config: dict,
    filepath: str | Path,
) -> Path:
    """
    Export results in a standardized two-sheet format for Power BI templates.

    Sheet 1 — 'Data': query results with formatted headers.
    Sheet 2 — 'ChartConfig': key/value pairs for chart configuration.

    Args:
        results_df: DataFrame of query results.
        viz_config: Visualization config dict from viz_recommender.
        filepath:   Destination .xlsx path.

    Returns:
        Path to the written file.
    """
    filepath = Path(filepath)
    filepath.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()

    # ── Sheet 1: Data ──────────────────────────────────────────
    ws_data = wb.active
    ws_data.title = "Data"

    if not results_df.empty:
        for col_idx, col_name in enumerate(results_df.columns, start=1):
            ws_data.cell(row=1, column=col_idx, value=col_name)
        for row_idx, row in enumerate(results_df.itertuples(index=False), start=2):
            for col_idx, value in enumerate(row, start=1):
                ws_data.cell(row=row_idx, column=col_idx, value=value)
        _style_header_row(ws_data, len(results_df.columns))
        _alternate_row_fill(ws_data, len(results_df), len(results_df.columns))
        _auto_column_width(ws_data)
        ws_data.freeze_panes = "A2"
    else:
        ws_data["A1"] = "No data returned."

    # ── Sheet 2: ChartConfig ───────────────────────────────────
    ws_cfg = wb.create_sheet("ChartConfig")
    config_rows = [
        ("chart_type", viz_config.get("chart_type", "")),
        ("x_axis",     viz_config.get("x_axis", "")),
        ("y_axis",     viz_config.get("y_axis", "")),
        ("color_by",   viz_config.get("color_by", "") or ""),
        ("title",      viz_config.get("title", "")),
        ("orientation", viz_config.get("orientation", "v")),
        ("powerbi_suggestion", viz_config.get("powerbi_suggestion", "")),
    ]
    ws_cfg.cell(row=1, column=1, value="Setting")
    ws_cfg.cell(row=1, column=2, value="Value")
    _style_header_row(ws_cfg, 2)
    for row_idx, (key, val) in enumerate(config_rows, start=2):
        ws_cfg.cell(row=row_idx, column=1, value=key)
        ws_cfg.cell(row=row_idx, column=2, value=str(val))
    _auto_column_width(ws_cfg)

    wb.save(filepath)
    return filepath


def export_standardized_excel_bytes(
    results_df: pd.DataFrame,
    viz_config: dict,
) -> bytes:
    """Return standardized Excel as bytes for Streamlit download."""
    buf = io.BytesIO()
    tmp_path = Path("_tmp_pbi_export.xlsx")
    export_standardized_excel(results_df, viz_config, tmp_path)
    with open(tmp_path, "rb") as f:
        data = f.read()
    tmp_path.unlink(missing_ok=True)
    return data


# ─────────────────────────────────────────────
# Power BI instructions generator
# ─────────────────────────────────────────────

def generate_powerbi_instructions(viz_config: dict) -> str:
    """
    Generate step-by-step Power BI Desktop setup instructions
    from a visualization config.

    Args:
        viz_config: Visualization config dict from viz_recommender.

    Returns:
        A formatted plain-text instruction string.
    """
    chart_type = viz_config.get("chart_type", "bar")
    x_axis = viz_config.get("x_axis", "")
    y_axis = viz_config.get("y_axis", "")
    color_by = viz_config.get("color_by") or ""
    title = viz_config.get("title", "")
    pbi_suggestion = viz_config.get("powerbi_suggestion", "")

    # Map chart types to Power BI visual names
    chart_map = {
        "bar":      "Clustered Bar Chart  OR  Clustered Column Chart",
        "line":     "Line Chart",
        "pie":      "Pie Chart  OR  Donut Chart",
        "scatter":  "Scatter Chart",
        "heatmap":  "Matrix Visual",
        "table":    "Table Visual",
        "kpi_card": "Card Visual  OR  KPI Visual",
    }
    pbi_visual = chart_map.get(chart_type, "Clustered Bar Chart")

    instructions = f"""
=============================================================
  POWER BI DASHBOARD SETUP GUIDE
  Generated by AI Dashboard Generator
=============================================================

REPORT TITLE : {title}
VISUAL TYPE  : {pbi_visual}

-------------------------------------------------------------
STEP 1 — LOAD THE DATA
-------------------------------------------------------------
1. Open Power BI Desktop
2. Click Home > Get Data > Excel Workbook
3. Select the exported file "powerbi_export.xlsx"
4. In the Navigator, select the "Data" sheet and click Load
5. The data will appear in the Fields pane on the right

-------------------------------------------------------------
STEP 2 — CREATE THE VISUAL
-------------------------------------------------------------
1. In the Visualizations pane, select: {pbi_visual}
2. Configure the fields:
   - Axis  / X-Axis: drag "{x_axis}" from the Fields pane
   - Values / Y-Axis: drag "{y_axis}" from the Fields pane
{f'   - Legend / Color:  drag "{color_by}" from the Fields pane' if color_by else '   - Legend: (not required for this chart)'}

-------------------------------------------------------------
STEP 3 — FORMAT THE VISUAL
-------------------------------------------------------------
1. Click the Format icon (paint roller) in Visualizations pane
2. Under "Title", set the title to: {title}
3. Under "Data labels", toggle ON to show values on bars/points
4. Under "Colors", apply your brand color scheme
5. Adjust font sizes: Title 14pt, Axis labels 11pt

-------------------------------------------------------------
STEP 4 — ADDITIONAL RECOMMENDATION
-------------------------------------------------------------
{pbi_suggestion}

-------------------------------------------------------------
STEP 5 — PUBLISH (optional)
-------------------------------------------------------------
1. Click Home > Publish
2. Select your Power BI workspace
3. Open Power BI Service in your browser to view the report

-------------------------------------------------------------
NOTE ON MCP SERVER AUTOMATION
-------------------------------------------------------------
If you have the Power BI Modeling MCP server configured with
Claude Desktop / VS Code + GitHub Copilot, you can automate
all of the above using the MCP prompts in the MCP Command
Generator panel of this app.

=============================================================
"""
    return instructions.strip()


# ─────────────────────────────────────────────
# DAX measure generator
# ─────────────────────────────────────────────

def generate_dax_measures(viz_config: dict, user_question: str) -> list[dict]:
    """
    Ask Claude to generate useful DAX measures for the given visualization.

    Args:
        viz_config:     Visualization config dict from viz_recommender.
        user_question:  Original user question.

    Returns:
        List of dicts: [{"name": "...", "dax": "...", "explanation": "..."}]
    """
    x_axis = viz_config.get("x_axis", "")
    y_axis = viz_config.get("y_axis", "")
    chart_type = viz_config.get("chart_type", "bar")
    title = viz_config.get("title", "")

    prompt = f"""You are a Power BI DAX expert. A user has asked this question:
"{user_question}"

The resulting chart is a {chart_type} with:
- X-axis / Dimension: {x_axis}
- Y-axis / Measure:   {y_axis}
- Chart title:        {title}

Generate 3-5 useful DAX measures that would enhance this analysis in Power BI.
Consider: YoY growth, MoM growth, % of total, running totals, rankings, or moving averages —
whichever are most relevant to the chart above.

Respond with a JSON array (no markdown, no code fences) with this structure:
[
  {{
    "name": "Total Revenue",
    "dax": "Total Revenue = SUM('Data'[{y_axis}])",
    "explanation": "Sums all revenue values from the data table."
  }},
  ...
]

Rules:
- Table name in DAX should be 'Data' (matching the Excel sheet name)
- Use the exact column name '{y_axis}' for the measure column
- Make each measure genuinely useful and different from the others
- Keep DAX syntax correct and Power BI compatible
"""

    client = anthropic.Anthropic(api_key=ANTHROPIC_API_KEY)
    message = client.messages.create(
        model=CLAUDE_MODEL,
        max_tokens=1024,
        messages=[{"role": "user", "content": prompt}],
    )

    raw = message.content[0].text.strip()
    if raw.startswith("```"):
        raw = "\n".join(
            line for line in raw.split("\n") if not line.startswith("```")
        ).strip()

    try:
        measures = json.loads(raw)
        if not isinstance(measures, list):
            measures = []
    except json.JSONDecodeError:
        measures = []

    return measures


# ─────────────────────────────────────────────
# MCP command generator
# ─────────────────────────────────────────────

def generate_mcp_commands(
    viz_config: dict,
    dax_measures: list[dict],
    results_df: pd.DataFrame,
    user_question: str,
) -> str:
    """
    Generate natural language prompts that a user can paste into Claude Desktop
    to trigger the Power BI Modeling MCP server.

    Args:
        viz_config:    Visualization config dict from viz_recommender.
        dax_measures:  List of DAX measures from generate_dax_measures().
        results_df:    DataFrame of query results.
        user_question: Original user question.

    Returns:
        A formatted string with MCP command prompts.
    """
    title = viz_config.get("title", "My Dashboard")
    x_axis = viz_config.get("x_axis", "")
    y_axis = viz_config.get("y_axis", "")
    chart_type = viz_config.get("chart_type", "bar")
    columns = list(results_df.columns) if not results_df.empty else [x_axis, y_axis]

    # Build column type hints
    col_types = []
    if not results_df.empty:
        for col in results_df.columns:
            dtype = str(results_df[col].dtype)
            if "int" in dtype or "float" in dtype:
                pbi_type = "decimal number"
            elif "datetime" in dtype:
                pbi_type = "date/time"
            else:
                pbi_type = "text"
            col_types.append(f"{col} ({pbi_type})")
    else:
        col_types = columns

    # Build DAX section
    dax_lines = []
    for m in dax_measures:
        dax_lines.append(f'        Add a measure \'{m["name"]}\' = {m["dax"]}')
    dax_block = "\n".join(dax_lines) if dax_lines else "        (no DAX measures)"

    # Map chart type to Power BI visual
    visual_map = {
        "bar":      "Clustered column chart",
        "line":     "Line chart",
        "pie":      "Donut chart",
        "scatter":  "Scatter chart",
        "heatmap":  "Matrix",
        "table":    "Table",
        "kpi_card": "Card",
    }
    pbi_visual = visual_map.get(chart_type, "Clustered column chart")

    col_types_str = ", ".join(col_types)

    return f"""
=============================================================
  MCP COMMAND PROMPTS FOR CLAUDE DESKTOP
  (Requires Power BI Modeling MCP Server)
=============================================================

Copy the prompt below and paste it into Claude Desktop.
Make sure Power BI Desktop is open with your target report.

-------------------------------------------------------------
PROMPT 1 — CREATE TABLE & LOAD DATA
-------------------------------------------------------------
Connect to Power BI Desktop. Create a new table called 'Data'
with the following columns: {col_types_str}.
Load the data from the file 'powerbi_export.xlsx', sheet 'Data'.

-------------------------------------------------------------
PROMPT 2 — CREATE DAX MEASURES
-------------------------------------------------------------
In the Power BI Desktop report, in the 'Data' table:
{dax_block}

-------------------------------------------------------------
PROMPT 3 — CREATE VISUAL
-------------------------------------------------------------
Add a {pbi_visual} to the report page.
Set the title to '{title}'.
Set the axis field to '{x_axis}'.
Set the values field to '{y_axis}'.
Format the visual with data labels enabled and a clean white background.

-------------------------------------------------------------
ORIGINAL USER QUESTION (for context)
-------------------------------------------------------------
"{user_question}"

=============================================================
""".strip()
